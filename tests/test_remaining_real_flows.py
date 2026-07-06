"""Acceptance coverage for the remaining real PostgreSQL-bound flows."""

from __future__ import annotations

import copy
import json
from pathlib import Path

import pytest

from up_to_postgresql.config.resolver import resolve_flow_config
from up_to_postgresql.config.schema import FlowConfig
from up_to_postgresql.flows import run_flow
from up_to_postgresql.readers import read_source


FLOW_CASES = {
    "llee_trams_demora": {
        "source": "llee_mensual.xlsx",
        "sheet": "Trams demora per Centre i Prova",
        "shape": (473, 15),
        "target_table": "llee_trams_demora",
        "load_mode": "append",
        "columns": [
            "any",
            "mes",
            "Centre",
            "Prioritat",
            "Grup de monitorització",
            "Total pacients",
            "porc_oportunidad",
            "0-30 d",
            "31-60 d",
            "61-90 d",
            "91-120 d",
            "121-150 d",
            "151-180 d",
            "181-365 d",
            "Mes d'1 any",
        ],
        "duplicate_key": [
            "any",
            "mes",
            "Centre",
            "Prioritat",
            "Grup de monitorització",
        ],
        "duplicate_rows": 0,
        "targets": [
            "any_llee",
            "mes_llee",
            "centre",
            "prioritat",
            "grup_monitoritzacio",
            "total_pacients",
            "percentatge_oportunitat",
            "tram_0_30_dies",
            "tram_31_60_dies",
            "tram_61_90_dies",
            "tram_91_120_dies",
            "tram_121_150_dies",
            "tram_151_180_dies",
            "tram_181_365_dies",
            "mes_1_any",
        ],
    },
    "td_ambulatoris": {
        "source": "td.xlsx",
        "sheet": "td_ambulatoris",
        "shape": (12996, 17),
        "target_table": "td_ambulatoris",
        "load_mode": "append",
        "columns": [
            "any",
            "mes",
            "centre_realitzacio",
            "sala_realitzacio",
            "data_realitzacio",
            "dies_previstos",
            "data_conveni",
            "prestacio_modalitat",
            "pte_num",
            "prestacio_denominacio",
            "especialista",
            "situacio",
            "data_resultats",
            "pte_cip",
            "data_assignacio",
            "dias_trigats",
            "dias_retard",
        ],
        "duplicate_key": [],
        "duplicate_rows": 492,
        "targets": [
            "any_ambulatoris",
            "mes_ambulatoris",
            "centre_realitzacio",
            "sala_realitzacio",
            "data_realitzacio",
            "dies_previstos",
            "data_conveni",
            "prestacio_modalitat",
            "pacient_numero",
            "prestacio_denominacio",
            "especialista",
            "situacio",
            "data_resultats",
            "pacient_cip",
            "data_assignacio",
            "dies_trigats",
            "dies_retard",
        ],
    },
    "td_urgencies": {
        "source": "td.xlsx",
        "sheet": "td_urgencies",
        "shape": (588, 6),
        "source_shape": (588, 8),
        "target_table": "td_urgencies",
        "load_mode": "append",
        "columns": [
            "solicitant",
            "any",
            "mes",
            "episodi",
            "prestació",
            "proves",
        ],
        "source_columns": [
            "solicitant",
            "any",
            "mes",
            "episodi",
            "prestació",
            "proves",
            "lat",
            "lon",
        ],
        "duplicate_key": ["solicitant", "any", "mes", "episodi", "prestació"],
        "duplicate_rows": 0,
        "targets": [
            "solicitant",
            "any_urgencies",
            "mes_urgencies",
            "episodi",
            "prestacio",
            "proves",
        ],
    },
    "demanda": {
        "source": "demanda._samplexlsx.xlsx",
        "sheet": "Demanda",
        "shape": (9, 23),
        "target_table": "demanda",
        "load_mode": "replace_partition",
        "partition_column": "any_prestacio",
        "columns": [
            "Any prestació (YYYY)",
            "Mes prestació (MM)",
            "Data prestació",
            "Número de prestacions",
            "Centre SAP codi (Hospital)",
            "Centre Codi",
            "Institució",
            "Institució corregida",
            "Màquina de la prestació codi",
            "Màquina de la prestació desc",
            "Màquina Descripció",
            "Prestació descripció",
            "Prestació nivell 9 codi",
            "Prestació nivell 9 desc",
            "Nivell 1 codi",
            "Nivell 1 desc",
            "UP sol·licitant desc",
            "UP sol·licitant entitat proveïdora desc",
            "UP Sol·licitant Corregida",
            "UP Sol·licitant Entitat Corregida",
            "UT sol·licitant desc",
            "Prestació estat codi",
            "Pacient (NHC)",
        ],
        "duplicate_key": [
            "Data prestació",
            "Prestació nivell 9 codi",
            "Pacient (NHC)",
        ],
        "duplicate_rows": 0,
        "targets": [
            "any_prestacio",
            "mes_prestacio",
            "data_prestacio",
            "numero_prestacions",
            "centre_sap_codi",
            "centre_codi",
            "institucio",
            "institucio_corregida",
            "maquina_prestacio_codi",
            "maquina_prestacio_descripcio",
            "maquina_descripcio",
            "prestacio_descripcio",
            "prestacio_nivell_9_codi",
            "prestacio_nivell_9_descripcio",
            "nivell_1_codi",
            "nivell_1_descripcio",
            "up_solicitant_descripcio",
            "up_solicitant_entitat_proveidora_descripcio",
            "up_solicitant_corregida",
            "up_solicitant_entitat_corregida",
            "ut_solicitant_descripcio",
            "prestacio_estat_codi",
            "pacient_nhc",
        ],
    },
}


DEMANDA_SOURCE_COLUMNS = [
    "Any prestació (YYYY)",
    "Mes prestació (MM)",
    "Data prestació",
    "Número de prestacions",
    "Centre SAP codi (Hospital)",
    "Centre Codi",
    "Institució",
    "Institució corregida",
    "Màquina de la prestació codi",
    "Màquina de la prestació desc",
    "Màquina Descripció",
    "Prestació descripció",
    "Prestació nivell 9 codi",
    "Prestació nivell 9 desc",
    "Nivell 1 codi",
    "Nivell 1 desc",
    "UP sol·licitant desc",
    "UP sol·licitant entitat proveïdora desc",
    "UP Sol·licitant Corregida",
    "UP Sol·licitant Entitat Corregida",
    "UT sol·licitant desc",
    "Prestació estat codi",
    "Pacient (NHC)",
]


def _build_demanda_xlsx(tmp_path: Path) -> Path:
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    source = tmp_path / "demanda_contract.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Demanda"

    headers = [None, *DEMANDA_SOURCE_COLUMNS]
    row = [
        None,
        "2026",
        "01",
        "2026-01-15",
        "1",
        "H01",
        "C01",
        "Institució A",
        "Institució A corregida",
        "MAC-01",
        "Màquina A",
        "Màquina A desc",
        "Prestació A",
        "P9-001",
        "Prestació nivell 9 desc A",
        "N1-001",
        "Nivell 1 desc A",
        "UP sol·licitant A",
        "UP sol·licitant entitat proveïdora A",
        "UP Sol·licitant Corregida A",
        "UP Sol·licitant Entitat Corregida A",
        "UT sol·licitant A",
        "ESTAT",
        "NHC-001",
    ]

    for column_index, value in enumerate(headers, start=1):
        sheet.cell(row=2, column=column_index, value=value)
    for column_index, value in enumerate(row, start=1):
        sheet.cell(row=3, column=column_index, value=value)

    workbook.save(source)
    return source


@pytest.mark.parametrize("flow_name", FLOW_CASES)
def test_remaining_real_flow_config_contract(flow_name: str) -> None:
    case = FLOW_CASES[flow_name]
    config = resolve_flow_config(flow_name, "test")
    expected_header_row = 2 if flow_name == "demanda" else 1

    assert config.data["source"] == {
        "encoding": "utf-8",
        "delimiter": ",",
        "type": "xlsx",
        "path": case["source"],
        "sheet": case["sheet"],
        "header_row": expected_header_row,
    }
    assert config.data["processing"] == {
        "drop_empty_rows": True,
        "drop_empty_columns": True,
        "trim_strings": "all",
    }
    assert config.data["validation"]["required_columns"] == case["columns"]
    assert config.data["validation"].get("duplicate_key", []) == case["duplicate_key"]
    assert config.data["validation"]["duplicate_policy"] == "report"
    assert config.data["validation"]["missing_columns_policy"] == "error"
    assert config.data["load"]["target_table"] == case["target_table"]
    assert config.data["load"]["load_mode"] == case["load_mode"]
    if "partition_column" in case:
        assert config.data["load"]["partition_column"] == case["partition_column"]
    else:
        assert "partition_column" not in config.data["load"]
    assert [item["source"] for item in config.data["load"]["column_mapping"]] == case[
        "columns"
    ]
    assert [item["target"] for item in config.data["load"]["column_mapping"]] == case[
        "targets"
    ]


@pytest.mark.parametrize("flow_name", FLOW_CASES)
def test_remaining_real_flow_reads_expected_xlsx_columns(
    flow_name: str, tmp_path: Path
) -> None:
    pytest.importorskip("pandas")
    pytest.importorskip("openpyxl")
    case = FLOW_CASES[flow_name]
    config = resolve_flow_config(flow_name, "test")

    if flow_name == "demanda":
        source = _build_demanda_xlsx(tmp_path)
        data = copy.deepcopy(config.data)
        data["paths"]["input_base_dir"] = str(tmp_path)
        data["source"] = {
            "type": "xlsx",
            "path": source.name,
            "sheet": "Demanda",
            "header_row": 2,
        }
        config = FlowConfig(name=config.name, env=config.env, data=data)
        source_shape = (1, 24)
        source_columns = ["Unnamed: 0", *DEMANDA_SOURCE_COLUMNS]
    else:
        source_shape = case.get("source_shape", case["shape"])
        source_columns = case.get("source_columns", case["columns"])
        source = Path("data/input/test") / str(case["source"])
        assert source.exists(), f"Missing real fixture: {source}"

    frame = read_source(config)

    assert frame.shape == source_shape
    assert list(frame.columns) == source_columns


@pytest.mark.parametrize("flow_name", FLOW_CASES)
def test_remaining_real_flow_writes_outputs_without_loading(
    flow_name: str, tmp_path: Path
) -> None:
    pytest.importorskip("pandas")
    pytest.importorskip("openpyxl")
    case = FLOW_CASES[flow_name]
    source_shape = case.get("source_shape", case["shape"])
    rows_written = case["shape"][0]
    columns_written = case["shape"][1]
    config = resolve_flow_config(flow_name, "test")
    if flow_name == "demanda":
        source = _build_demanda_xlsx(tmp_path)
        data = copy.deepcopy(config.data)
        data["paths"]["input_base_dir"] = str(tmp_path)
        data["paths"]["output_base_dir"] = str(tmp_path / "output")
        data["source"] = {
            "type": "xlsx",
            "path": source.name,
            "sheet": "Demanda",
            "header_row": 2,
        }
        config = FlowConfig(name=config.name, env=config.env, data=data)
        source_shape = (1, 24)
        rows_written = 1
        columns_written = 23
    config.data["paths"]["output_base_dir"] = str(tmp_path)

    result = run_flow(config)

    assert result.rows_read == source_shape[0]
    assert result.columns_read == source_shape[1]
    assert result.rows_written == rows_written
    assert result.columns_written == columns_written
    assert result.empty_rows_removed == 0
    if flow_name == "demanda":
        assert result.empty_columns_removed == ("Unnamed: 0",)
    else:
        assert result.empty_columns_removed == tuple(case.get("empty_columns_removed", []))
    assert result.duplicate_rows == case["duplicate_rows"]
    assert result.processed_path.exists()
    assert result.report_path.exists()

    report = json.loads(result.report_path.read_text(encoding="utf-8"))
    assert report["duplicate_key"] == case["duplicate_key"]
    assert report["duplicate_rows"] == case["duplicate_rows"]
    assert report["postgresql"] == {"status": "skipped"}
    if flow_name == "demanda":
        assert result.empty_columns_removed == ("Unnamed: 0",)
        processed = result.processed_path.read_text(encoding="utf-8").splitlines()
        assert processed[0].split(",") == DEMANDA_SOURCE_COLUMNS
        assert processed[1].split(",")[0] == "2026"


@pytest.mark.parametrize(
    ("filename", "content"),
    [
        (
            "td_urgencies_with_empty_lat_lon.csv",
            (
                "solicitant,any,mes,episodi,prestació,proves,lat,lon\n"
                "Hospital A,2026,05,E-1,RX,2,,\n"
            ),
        ),
        (
            "td_urgencies_without_lat_lon.csv",
            (
                "solicitant,any,mes,episodi,prestació,proves\n"
                "Hospital A,2026,05,E-1,RX,2\n"
            ),
        ),
    ],
)
def test_td_urgencies_processes_sources_with_or_without_lat_lon(
    tmp_path: Path, filename: str, content: str
) -> None:
    pytest.importorskip("pandas")
    source = tmp_path / filename
    source.write_text(content, encoding="utf-8")
    resolved = resolve_flow_config("td_urgencies", "test")
    data = copy.deepcopy(resolved.data)
    data["paths"]["input_base_dir"] = str(tmp_path)
    data["paths"]["output_base_dir"] = str(tmp_path / "output")
    data["source"] = {
        "type": "csv",
        "path": filename,
        "encoding": "utf-8",
        "delimiter": ",",
    }
    config = FlowConfig(name=resolved.name, env=resolved.env, data=data)

    result = run_flow(config)

    assert result.missing_required_columns == ()
    assert result.rows_written == 1
    assert result.columns_written == 6
    assert result.processed_path.read_text(encoding="utf-8").splitlines() == [
        "solicitant,any,mes,episodi,prestació,proves",
        "Hospital A,2026,05,E-1,RX,2",
    ]
    assert [item["source"] for item in config.data["load"]["column_mapping"]] == [
        "solicitant",
        "any",
        "mes",
        "episodi",
        "prestació",
        "proves",
    ]
